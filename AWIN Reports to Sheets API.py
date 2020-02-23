# Authentication: Use setup.xlsx

# Set to True to display API responses for debugging.
include_response = True

import datetime
import requests 
import json
import openpyxl
from time import sleep
import pickle
import os.path
from googleapiclient import discovery
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from oauth2client.service_account import ServiceAccountCredentials

# Google Sheets Auth
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
TOKEN_FILE = 'token.pickle'
CREDENTIAL_FILE = 'credentials.json'
credentials = None

if os.path.exists(TOKEN_FILE):
    with open(TOKEN_FILE, 'rb') as token:
        credentials = pickle.load(token)

if not credentials or not credentials.valid:
    if credentials and credentials.expired and credentials.refresh_token:
        credentials.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(
            CREDENTIAL_FILE, SCOPES)
        credentials = flow.run_local_server(port=10800)
    # Save the credentials for the next run
    with open(TOKEN_FILE, 'wb') as token:
        pickle.dump(credentials, token)

service = discovery.build('sheets', 'v4', credentials=credentials)

#########################################################################################################

def exportreport(AdvertiserName, AdvertiserID, spreadsheet_ID):
    print("Checking for open sales") 
    updatecounter = 0
    ranges = "Assistant Sheet!B:B"
    include_grid_data = True
    requestopen = service.spreadsheets().get(spreadsheetId=spreadsheet_ID, ranges=ranges, includeGridData=include_grid_data)
    responseopen = requestopen.execute() # Requests a list of addresses (A1-notation) of open sales, created using the ADDRESS formula on the Assistant Sheet.
    if include_response == True:
        print("Google Sheets response: " + str(responseopen))
        
    for i in range(0,4000):
        if "formattedValue" in str(responseopen["sheets"][0]["data"][0]["rowData"][i]):
            opensales = responseopen["sheets"][0]["data"][0]["rowData"][i]["values"][0]["formattedValue"]
            rownumber = opensales[3]
            for m in range(4,8):
                try:
                    rownumber += opensales[m]
                except IndexError:
                    continue

            range_open = "Input Awin Publisher!A" + str(rownumber) + ":F" + str(rownumber)

            requestopen2 = service.spreadsheets().values().get(spreadsheetId=spreadsheet_ID, range=range_open)
            responseopen2 = requestopen2.execute()
            if include_response == True:
                print("Open Sale found: " + str(responseopen2))

            end_dateopen = responseopen2["values"][0][0]
            start_dateopenlist = []
            for i in range(0,8):
                start_dateopenlist += responseopen2["values"][0][0][i]
            start_dateopenlist += "01"
            start_dateopen = "".join(start_dateopenlist)
            publisher_ID = responseopen2["values"][0][2]

            urlopenawin = "https://api.awin.com/advertisers/" + str(AdvertiserID) + "/reports/publisher?startDate=" + str(start_dateopen) + "&endDate=" + str(end_dateopen) + "&timezone=Europe/Berlin&accessToken=" + str(AwinToken)
            payloadopenawin = {}
            responseopenawin = requests.request("GET", urlopenawin, data = payloadopenawin) # Requests all sales of the month that contains open sales.

            if include_response == True:
                print("AWIN response: " + str(responseopenawin))
            if "401" in str(responseopenawin):
                print("401: Failed to auth. Please check the Awin OAuth2 Token inside the 'Setup.xlsx' and update it if necessary.")
            outputopenawin = json.loads(responseopenawin.text.encode('utf8'))
            lenawin = len(outputopenawin)

            for k in range (0, lenawin):
                x = outputopenawin[k]["publisherId"]
                y = publisher_ID
                if int(x) == int(y):
                    rangeupdatesales = "Input Awin Publisher!F" + str(rownumber) + ":T" + str(rownumber)
                    
                    try:
                        updatepayload = [int(outputopenawin[k]['pendingNo']), int(outputopenawin[k]['pendingValue']), int(outputopenawin[k]['pendingComm']), int(outputopenawin[k]['confirmedNo']), int(outputopenawin[k]['confirmedValue']), int(outputopenawin[k]['confirmedComm']), int(outputopenawin[k]['bonusNo']), int(outputopenawin[k]['bonusValue']), int(outputopenawin[k]['bonusComm']), int(outputopenawin[k]['totalNo']), int(outputopenawin[k]['totalValue']), int(outputopenawin[k]['totalComm']), int(outputopenawin[k]['declinedNo']), int(outputopenawin[k]['declinedValue']), int(outputopenawin[k]['declinedComm'])]
                    except KeyError:
                        print("KeyError")
                        print("DEBUG: " + str(updatepayload))
                    
                    value_input_option = 'USER_ENTERED'  
                    value_range_body = {
                                        "range": rangeupdatesales,
                                        "values": [
                                         updatepayload
                                        ]                                        
                                        }

                    requestupdatesales = service.spreadsheets().values().update(spreadsheetId=spreadsheet_ID, range=rangeupdatesales, valueInputOption=value_input_option, body=value_range_body)
                    responseupdatesales = requestupdatesales.execute() # Updates monthly statistic of the respective publisher.
                    updatecounter += + 1
                    if include_response == True:
                        print("Google Sheets response: " + str(responseupdatesales))
                    print("Open sale updated.")
                    countdown()
                    
                else:
                    continue
        else:
            continue
    if updatecounter == 0:
        print("No open sales found.")
    else:
        print("Successfully updated " + str(updatecounter) + " open sales for " + str(AdvertiserName) + ".")
    
    # Export previous month from Awin
    # Checks, whether previous month has already been imported or not.

    print("Checking for previous imports") 
    rangelastm = "Input Awin Publisher!A:A"
    requestlastm = service.spreadsheets().values().get(spreadsheetId=spreadsheet_ID, range=rangelastm, majorDimension="ROWS")    
    responselastm = requestlastm.execute() # Requests a list of months already in the sheets, to check whether the previous month has already been imported.
    if include_response == True:
        print("Google Sheets response: " + str(responselastm))
    listofdate = [str(end_date)]
    try:
        if listofdate in responselastm["values"]:
            print("Last month already imported. Skipping import.")
            return(1)
    except:
        pass
    print("Last month not yet imported. Starting import.")
    url = "https://api.awin.com/advertisers/" + str(AdvertiserID) + "/reports/publisher?startDate=" + str(start_date) + "&endDate=" + str(end_date) + "&timezone=Europe/Berlin&accessToken=" + str(AwinToken)
    payload = {}

    response = requests.request("GET", url, data = payload)
    if include_response == True:
        print("AWIN response: " + str(response))
    if "401" in str(response):
        print("401: Failed to auth. Please check the Awin OAuth2 Token inside the 'Setup.xlsx' and update it if necessary.")
    output = json.loads(response.text.encode('utf8'))
    length = len(output)
    print(str(AdvertiserName) + ": Export done.")
    
    # Import to Google Sheets
    # Adds data starting from the first empty row.
    
    columns = []
    for i in range(0, length):
        try:
            columns += [[str(end_date),output[i]['publisherName'],output[i]['publisherId'],output[i]['impressions'],output[i]['clicks'],output[i]['pendingNo'],output[i]['pendingValue'],output[i]['pendingComm'],output[i]['confirmedNo'],output[i]['confirmedValue'],output[i]['confirmedComm'],output[i]['totalNo'],output[i]['totalValue'],output[i]['totalComm'],output[i]['declinedNo'],output[i]['declinedValue'],output[i]['declinedComm'],output[i]['bonusNo'],output[i]['bonusValue'],output[i]['bonusComm']]]
        except KeyError:
            print("Error: KeyError for " + str(AdvertiserName) + ". Either no values yet or failed to auth.")

    if rowcount == 0: # Adds headers if they do not exist yet.
        range_3 = "Input Awin Publisher!A1:T1"
        value_input_option = "RAW"
        value_range_body = {
            "values": [
                ["Monat", "Publisher", "Publisher-ID", "Views", "Alle_Klicks", "Offen_Anz", "Gesamt_offener_Umsatz", "Offene_Prov", "Bestätigte_Anz", "Gesamt_bestätigter_Umsatz", "Bestätigte_Prov", "Bonus_Anz", "Bonus_Umsatz", "Bonus_Prov", "Gesamt_Anz", "Gesamt_Umsatz", "Gesamt_Prov", "Abgelehnt_Anz", "Abgelehnter_Umsatz", "Abgelehnt_Prov"]
            ]
        }
        requestheader = service.spreadsheets().values().update(spreadsheetId=spreadsheet_ID, range=range_3, body=value_range_body, valueInputOption = value_input_option)
        responseheader = requestheader.execute()
        if include_response == True:
            print("Google Sheets response: " + str(responseheader))
        rowcountlocal = rowcount + 1
        print("Column headers added.")

    else:
        rowcountlocal = rowcount 

    range_X = 'Input Awin Publisher!A2:T'  
    value_input_option = 'RAW'
    value_range_body = {

      "values": 
        columns   
    }
    requestupdate = service.spreadsheets().values().update(spreadsheetId=spreadsheet_ID, range=range_X, valueInputOption=value_input_option, body=value_range_body)
    responseupdate = requestupdate.execute()
    if include_response == True:
        print("Google Sheets response: " + str(responseupdate))
    print("Import " + str(AdvertiserName) + ": import done.")

## Creates new sheets, if they do not exist yet.
def checkingsheets(spreadsheet_ID, AdvertiserName):
    print("Checking Sheets")
    # Creates the Assistant Sheet
    try:
        ranges = "Assistant Sheet!B:B"
        include_grid_data = False
        requestsheetnames = service.spreadsheets().get(spreadsheetId=spreadsheet_ID, ranges=ranges, includeGridData=include_grid_data)
        responsesheetnames = requestsheetnames.execute()
        if include_response == True:
            print("Google Sheets response: " + str(responsesheetnames))
        print("Assistant Sheet found.")
    except:
        print("Assistant Sheet not found for " + str(AdvertiserName) + ".")
        batch_update_spreadsheet_request_body4 = {
         "requests": [
          {
           "addSheet": {
            "properties": {
                "title": "Assistant Sheet",
                "tabColor": {
                  "red": 1,
                  "green": 0,
                  "blue": 0,
                  "alpha": 1
                },
                },
            }
            }
          ],
          "includeSpreadsheetInResponse": False,
          "responseRanges": [
            
          ],
          "responseIncludeGridData": False
        }

        requestcreatesheet2 = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_ID, body=batch_update_spreadsheet_request_body4)
        responsecreatesheet2 = requestcreatesheet2.execute()
        if include_response == True:
            print("Google Sheets response: " + str(responsecreatesheet2))

        formula = []
        for i in range(2,10000):
            formula += [str("=IFERROR(IF('Input Awin Publisher'!F") + str(i) + str("=0;"";ADDRESS(") + str(i) + str(",6));"")")]
        print("formula: " + str(formula))
        
        batch_update_values_request_body6 = {
        'value_input_option': 'USER_ENTERED',  
        'data': [
                {
            "range": "Assistant Sheet!B2:B",
            "majorDimension": "COLUMNS",
            "values": [
                formula
                      ]
                }
                ],  
        }
        requestfillsheet = service.spreadsheets().values().batchUpdate(spreadsheetId=spreadsheet_ID, body=batch_update_values_request_body6)
        responsefillsheet = requestfillsheet.execute()
        
        print("Created Assistant Sheet for " + str(AdvertiserName) + ".")

    try:
        ranges = "Input Awin Publisher!A:A"
        include_grid_data = False
        requestsheetnames = service.spreadsheets().get(spreadsheetId=spreadsheet_ID, ranges=ranges, includeGridData=include_grid_data)
        responsesheetnames = requestsheetnames.execute()
        if include_response == True:
            print("Google Sheets response: " + str(responsesheetnames))
        sheet_id = responsesheetnames["sheets"][0]["properties"]["sheetId"]
        print("Sheet 'Input Awin Publisher' found")
        return(sheet_id)
    except:
        print("Sheet 'Input Awin Publisher' not found for " + str(AdvertiserName) + ".")
        batch_update_spreadsheet_request_body3 = {
         "requests": [
          {
           "addSheet": {
            "properties": {
                "title": "Input Awin Publisher",
                "tabColor": {
                  "red": 1,
                  "green": 0,
                  "blue": 0,
                  "alpha": 1
                },
                },
            }
            }
          ],
          "includeSpreadsheetInResponse": False,
          "responseRanges": [
            
          ],
          "responseIncludeGridData": False
        }

        requestcreatesheet = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_ID, body=batch_update_spreadsheet_request_body3)
        responsecreatesheet = requestcreatesheet.execute()
        if include_response == True:
            print("Google Sheets response: " + str(responsecreatesheet))
        sheet_id = responsecreatesheet["replies"][0]["addSheet"]["properties"]["sheetId"]
        print("Created Sheet 'Input Awin Publisher' for " + str(AdvertiserName) + ".")
        return(sheet_id)

    # Exports column A to determine the current number of rows, so that new data can be appended to the current data
def checkrows(spreadsheed_ID):
    range_1 = "Input Awin Publisher!A:A"
    requestrows = service.spreadsheets().values().get(spreadsheetId=spreadsheet_ID, range=range_1)
    rows = requestrows.execute()
    try:
        rowcount = len(rows["values"])
        return(rowcount)
    except KeyError: # A KeyError means there are no values yet, not even headers. Therefore, headers will be created later.
        print("No column headers. Adding column headers.")
        rowcount = 0
        return(rowcount)

    # A short countdown between exports. This is necessary as AWIN has a limit of 20 requests per minute.
def countdown():
    print("Starting next export in...")
    print("3")
    sleep(3)
    print("2")
    sleep(2)
    print("1")
    sleep(1)    

def main():
    # Read Setup.xlsx
    path = ("Setup.xlsx")
    wb = openpyxl.load_workbook(path)

    # Get previous month
    today = datetime.date.today()
    first = today.replace(day=1)
    end_date = first - datetime.timedelta(days=1)
    start_date = end_date.replace(day=1)

    # Retrieve Awin Access Token
    sheet1 = wb["Token"]
    cell2 = sheet1.cell(row = 2, column = 1)
    AwinToken = cell2.value
    print("AwinToken found.")

    # Retrieve list of programs and IDs
    sheet2 = wb["Setup"]
    max_row=sheet2.max_row
    max_column=sheet2.max_column

    for i in range(2, max_row+1):
        cell_AdvertiserID = sheet2.cell(row = i, column = 1)
        cell_AdvertiserName = sheet2.cell(row = i, column = 2)
        cell_spreadsheetID = sheet2.cell(row = i, column = 3)
        spreadsheet_ID = cell_spreadsheetID.value
        AdvertiserName = cell_AdvertiserName.value
        AdvertiserID = cell_AdvertiserID.value

        sheet_id = checkingsheets(spreadsheet_ID, AdvertiserName) # Checks if there is a Spreadsheet for each Advertiser. Also checks it there are the 2 needed sheets ("Assistant Sheet" and "Input Awin Publisher") and if necessary creates them.
        rowcount = checkrows(spreadsheet_ID) # Checks the number of rows and if headers already exist.

    # Exports data from the previous month

        exportreport(AdvertiserName, AdvertiserID, spreadsheet_ID)
        print(str(AdvertiserName) + ": All exports and imports done")
        countdown()
    ##    except:
    ##        print("Missing data in 'Setup.xlsx'. Skipping Advertiser.")
    ##        continue
        
    print("All programs done.")

if __name__ == "__main__":
    main()
